"""Excel出力: 分析結果をExcelファイルに書き出す"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analyzer import AnalysisResult
from parsers import ExtractedFinancials


# スタイル定義
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
_PRIORITY_FILLS = {
    "A": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "B": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "C": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
_RISK_FILLS = {
    "high": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
_BODY_FONT = Font(name="Yu Gothic", size=10)
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_style(ws, row, col, font=None, fill=None):
    cell = ws.cell(row=row, column=col)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = _WRAP_ALIGNMENT
    cell.border = _THIN_BORDER


def export_to_excel(
    result: AnalysisResult,
    extracted_financials: ExtractedFinancials | None = None,
) -> bytes:
    """分析結果をExcelバイト列として返す"""
    wb = Workbook()

    # ---- Sheet 1: 質問リスト ----
    ws1 = wb.active
    ws1.title = "質問リスト"

    headers1 = [
        ("No.", 6),
        ("カテゴリ", 22),
        ("優先度", 8),
        ("質問事項", 55),
        ("質問の意図・目的", 35),
        ("背景・根拠", 35),
        ("関連資料", 22),
        ("フォローアップ", 35),
        ("回答メモ", 30),
    ]

    for col_idx, (header, width) in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_ALIGNMENT
        cell.border = _THIN_BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    priority_order = {"A": 0, "B": 1, "C": 2}
    sorted_questions = sorted(
        result.questions,
        key=lambda q: (priority_order.get(q.priority, 3), q.category_code),
    )

    for i, q in enumerate(sorted_questions, 1):
        row = i + 1
        ws1.cell(row=row, column=1, value=i)
        ws1.cell(row=row, column=2, value=q.category_name)
        ws1.cell(row=row, column=3, value=q.priority)
        ws1.cell(row=row, column=4, value=q.question)
        ws1.cell(row=row, column=5, value=q.intent)
        ws1.cell(row=row, column=6, value=q.background)
        ws1.cell(row=row, column=7, value="\n".join(q.source_documents))
        ws1.cell(row=row, column=8, value="\n".join(q.follow_up_points))
        ws1.cell(row=row, column=9, value="")

        for col_idx in range(1, 10):
            _apply_style(ws1, row, col_idx, font=_BODY_FONT)

        fill = _PRIORITY_FILLS.get(q.priority)
        if fill:
            ws1.cell(row=row, column=3).fill = fill

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:I{len(sorted_questions) + 1}"

    # ---- Sheet 2: 主要論点・リスク ----
    ws2 = wb.create_sheet("主要論点・リスク")

    headers2 = [
        ("No.", 6),
        ("論点", 30),
        ("詳細", 55),
        ("リスクレベル", 12),
        ("M&Aリスクへの影響", 45),
        ("買収後の対策・解決策", 45),
        ("関連カテゴリ", 20),
    ]

    for col_idx, (header, width) in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_ALIGNMENT
        cell.border = _THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    for i, ki in enumerate(result.key_issues, 1):
        row = i + 1
        risk_label = {"high": "高", "medium": "中", "low": "低"}.get(ki.risk_level, ki.risk_level)
        cat_names = ", ".join(ki.related_categories)

        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=ki.title)
        ws2.cell(row=row, column=3, value=ki.description)
        ws2.cell(row=row, column=4, value=risk_label)
        ws2.cell(row=row, column=5, value=getattr(ki, "ma_risk_implications", ""))
        ws2.cell(row=row, column=6, value=getattr(ki, "post_merger_solutions", ""))
        ws2.cell(row=row, column=7, value=cat_names)

        for col_idx in range(1, 8):
            _apply_style(ws2, row, col_idx, font=_BODY_FONT)

        fill = _RISK_FILLS.get(ki.risk_level)
        if fill:
            ws2.cell(row=row, column=4).fill = fill

    ws2.freeze_panes = "A2"

    # ---- Sheet 3: 財務分析 ----
    ef = extracted_financials
    fa = result.financial_analysis
    has_csv = ef and (ef.pl_list or ef.bs_list)
    has_llm_pl = bool(fa.pl_trends)
    has_llm_bs = bool(fa.bs_trends)
    has_any = has_csv or has_llm_pl or has_llm_bs or fa.key_metrics or fa.financial_comments

    if has_any:
        ws_fin = wb.create_sheet("財務分析")
        fin_row = 1

        # --- 財務コメント (LLM) ---
        if fa.financial_comments:
            ws_fin.cell(row=fin_row, column=1, value="財務分析コメント")
            ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=11)
            fin_row += 1
            ws_fin.cell(row=fin_row, column=1, value=fa.financial_comments)
            ws_fin.cell(row=fin_row, column=1).font = _BODY_FONT
            ws_fin.cell(row=fin_row, column=1).alignment = _WRAP_ALIGNMENT
            ws_fin.merge_cells(start_row=fin_row, start_column=1, end_row=fin_row, end_column=8)
            fin_row += 2

        # --- 業績推移 (P/L) ---
        if has_csv and ef.pl_list:
            # CSV直接データ
            ws_fin.cell(row=fin_row, column=1, value=f"業績推移（P/L）　単位: 円　（{len(ef.pl_list)}期分・CSV抽出）")
            ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=11)
            fin_row += 1

            pl_headers = [
                ("期間", 22), ("売上高", 18), ("売上原価", 18), ("売上総利益", 18),
                ("販管費", 18), ("営業利益", 18), ("経常利益", 18), ("当期純利益", 18),
            ]
            for ci, (h, w) in enumerate(pl_headers, 1):
                cell = ws_fin.cell(row=fin_row, column=ci, value=h)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _WRAP_ALIGNMENT
                cell.border = _THIN_BORDER
                ws_fin.column_dimensions[get_column_letter(ci)].width = max(
                    ws_fin.column_dimensions[get_column_letter(ci)].width or 0, w
                )
            fin_row += 1

            for pl in ef.pl_list:
                ws_fin.cell(row=fin_row, column=1, value=pl.period_label)
                ws_fin.cell(row=fin_row, column=2, value=pl.revenue)
                ws_fin.cell(row=fin_row, column=3, value=pl.cost_of_sales)
                ws_fin.cell(row=fin_row, column=4, value=pl.gross_profit)
                ws_fin.cell(row=fin_row, column=5, value=pl.sga)
                ws_fin.cell(row=fin_row, column=6, value=pl.operating_profit)
                ws_fin.cell(row=fin_row, column=7, value=pl.ordinary_profit)
                ws_fin.cell(row=fin_row, column=8, value=pl.net_income)
                for ci in range(1, 9):
                    _apply_style(ws_fin, fin_row, ci, font=_BODY_FONT)
                    if ci >= 2:
                        ws_fin.cell(row=fin_row, column=ci).number_format = '#,##0'
                fin_row += 1
            fin_row += 1

        elif has_llm_pl:
            # LLMフォールバック（IM等から抽出）
            unit = fa.pl_trends[0].unit if fa.pl_trends else "千円"
            ws_fin.cell(row=fin_row, column=1, value=f"業績推移（P/L）　単位: {unit}　（{len(fa.pl_trends)}期分・IM抽出）")
            ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=11)
            fin_row += 1

            pl_headers = [
                ("期間", 22), ("売上高", 18), ("営業利益", 18),
                ("経常利益", 18), ("当期純利益", 18), ("EBITDA", 18),
            ]
            for ci, (h, w) in enumerate(pl_headers, 1):
                cell = ws_fin.cell(row=fin_row, column=ci, value=h)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _WRAP_ALIGNMENT
                cell.border = _THIN_BORDER
                ws_fin.column_dimensions[get_column_letter(ci)].width = max(
                    ws_fin.column_dimensions[get_column_letter(ci)].width or 0, w
                )
            fin_row += 1

            for p in fa.pl_trends:
                ws_fin.cell(row=fin_row, column=1, value=p.period)
                ws_fin.cell(row=fin_row, column=2, value=p.revenue)
                ws_fin.cell(row=fin_row, column=3, value=p.operating_profit)
                ws_fin.cell(row=fin_row, column=4, value=p.ordinary_profit)
                ws_fin.cell(row=fin_row, column=5, value=p.net_income)
                ws_fin.cell(row=fin_row, column=6, value=p.ebitda)
                for ci in range(1, 7):
                    _apply_style(ws_fin, fin_row, ci, font=_BODY_FONT)
                    if ci >= 2:
                        ws_fin.cell(row=fin_row, column=ci).number_format = '#,##0'
                fin_row += 1
            fin_row += 1

        # --- BS推移 ---
        if has_csv and ef.bs_list:
            ws_fin.cell(row=fin_row, column=1, value=f"BS推移（貸借対照表）　単位: 円　（{len(ef.bs_list)}期分・CSV抽出）")
            ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=11)
            fin_row += 1

            bs_headers = [
                ("期間", 22), ("総資産", 18), ("負債合計", 18),
                ("純資産", 18), ("現預金", 18), ("有利子負債", 18),
            ]
            for ci, (h, w) in enumerate(bs_headers, 1):
                cell = ws_fin.cell(row=fin_row, column=ci, value=h)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _WRAP_ALIGNMENT
                cell.border = _THIN_BORDER
            fin_row += 1

            for bs in ef.bs_list:
                ws_fin.cell(row=fin_row, column=1, value=bs.period_label)
                ws_fin.cell(row=fin_row, column=2, value=bs.total_assets)
                ws_fin.cell(row=fin_row, column=3, value=bs.total_liabilities)
                ws_fin.cell(row=fin_row, column=4, value=bs.net_assets)
                ws_fin.cell(row=fin_row, column=5, value=bs.cash)
                ws_fin.cell(row=fin_row, column=6, value=bs.interest_bearing_debt)
                for ci in range(1, 7):
                    _apply_style(ws_fin, fin_row, ci, font=_BODY_FONT)
                    if ci >= 2:
                        ws_fin.cell(row=fin_row, column=ci).number_format = '#,##0'
                fin_row += 1
            fin_row += 1

        elif has_llm_bs:
            # LLMフォールバック（IM等から抽出）
            unit = fa.bs_trends[0].unit if fa.bs_trends else "千円"
            ws_fin.cell(row=fin_row, column=1, value=f"BS推移（貸借対照表）　単位: {unit}　（{len(fa.bs_trends)}期分・IM抽出）")
            ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=11)
            fin_row += 1

            bs_headers = [
                ("期間", 22), ("総資産", 18), ("負債合計", 18),
                ("純資産", 18), ("現預金", 18), ("有利子負債", 18),
            ]
            for ci, (h, w) in enumerate(bs_headers, 1):
                cell = ws_fin.cell(row=fin_row, column=ci, value=h)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _WRAP_ALIGNMENT
                cell.border = _THIN_BORDER
            fin_row += 1

            for b in fa.bs_trends:
                ws_fin.cell(row=fin_row, column=1, value=b.period)
                ws_fin.cell(row=fin_row, column=2, value=b.total_assets)
                ws_fin.cell(row=fin_row, column=3, value=b.total_liabilities)
                ws_fin.cell(row=fin_row, column=4, value=b.net_assets)
                ws_fin.cell(row=fin_row, column=5, value=b.cash_and_deposits)
                ws_fin.cell(row=fin_row, column=6, value=b.interest_bearing_debt)
                for ci in range(1, 7):
                    _apply_style(ws_fin, fin_row, ci, font=_BODY_FONT)
                    if ci >= 2:
                        ws_fin.cell(row=fin_row, column=ci).number_format = '#,##0'
                fin_row += 1
            fin_row += 1

        # --- 主要財務指標 (LLM) ---
        if fa.key_metrics:
            ws_fin.cell(row=fin_row, column=1, value="主要財務指標")
            ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=11)
            fin_row += 1

            for km in fa.key_metrics:
                ws_fin.cell(row=fin_row, column=1, value=km.metric_name)
                ws_fin.cell(row=fin_row, column=1).font = Font(name="Yu Gothic", bold=True, size=10)
                ws_fin.cell(row=fin_row, column=1).border = _THIN_BORDER
                for vi, v in enumerate(km.values):
                    ci = vi + 2
                    cell = ws_fin.cell(row=fin_row, column=ci)
                    cell.value = f"{v.get('period', '')}: {v.get('value', '')}"
                    cell.font = _BODY_FONT
                    cell.alignment = _WRAP_ALIGNMENT
                    cell.border = _THIN_BORDER
                fin_row += 1
            fin_row += 1

        ws_fin.freeze_panes = "A2"

    # ---- Sheet 4: 企業サマリー ----
    ws3 = wb.create_sheet("企業サマリー")
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 80

    company_url = getattr(result, "company_url", "") or ""
    summary_data = [
        ("対象企業", result.company_name),
        ("企業URL", company_url),
        ("作成日", datetime.now().strftime("%Y年%m月%d日")),
        ("質問数", f"{len(result.questions)}問"),
        ("論点数", f"{len(result.key_issues)}件"),
        ("", ""),
        ("企業サマリー", result.summary),
    ]

    for i, (label, value) in enumerate(summary_data, 1):
        cell_a = ws3.cell(row=i, column=1, value=label)
        cell_a.font = Font(name="Yu Gothic", bold=True, size=10)
        cell_a.alignment = _WRAP_ALIGNMENT
        cell_b = ws3.cell(row=i, column=2, value=value)
        cell_b.font = _BODY_FONT
        cell_b.alignment = _WRAP_ALIGNMENT

    # バイト列で返す
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
